from io import BytesIO
from openpyxl import load_workbook

EXPECTED_HEADERS = [
    "ma_vach",
    "ten_san_pham",
    "danh_muc",
    "gia_ban",
    "link_anh",
    "tinh_trang_kho",
    "chien_luoc",
    "uu_tien",
]

VALID_TINH_TRANG = {"sẵn", "order"}
VALID_CHIEN_LUOC = {"mass", "thanh lý", "mở bán"}
VALID_UU_TIEN = {"1", "2", "3"}


def _normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def _parse_price(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        digits = "".join(c for c in text if c.isdigit() or c == ".")
        try:
            return float(digits) if digits else 0.0
        except ValueError:
            return 0.0


def parse_excel(file_bytes):
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("File Excel trống.")

    # Dòng 1: header, dòng 2: ghi chú, dòng 3 trở đi: dữ liệu
    headers = [_normalize(h).lower() for h in rows[0]]

    # Kiểm tra đúng định dạng chuẩn
    missing = [col for col in EXPECTED_HEADERS if col not in headers]
    if missing:
        raise ValueError(
            f"File không đúng định dạng chuẩn. Thiếu cột: {', '.join(missing)}. "
            "Vui lòng dùng file mẫu bang_san_pham_chuan.xlsx"
        )

    idx = {col: headers.index(col) for col in EXPECTED_HEADERS}

    products = []
    skipped = 0

    # Bỏ qua dòng 1 (header) và dòng 2 (ghi chú)
    for row in rows[2:]:
        if not row or all(cell is None or _normalize(cell) == "" for cell in row):
            continue

        name = _normalize(row[idx["ten_san_pham"]])
        if not name:
            continue

        ma_vach = _normalize(row[idx["ma_vach"]])
        danh_muc = _normalize(row[idx["danh_muc"]])
        gia_ban = _parse_price(row[idx["gia_ban"]])
        link_anh = _normalize(row[idx["link_anh"]])
        tinh_trang = _normalize(row[idx["tinh_trang_kho"]]).lower()
        chien_luoc = _normalize(row[idx["chien_luoc"]]).lower()
        uu_tien_raw = _normalize(row[idx["uu_tien"]])
        uu_tien = uu_tien_raw if uu_tien_raw in VALID_UU_TIEN else "2"

        # Chuẩn hoá giá trị
        if tinh_trang not in VALID_TINH_TRANG:
            tinh_trang = "sẵn"
        if chien_luoc not in VALID_CHIEN_LUOC:
            chien_luoc = "mass"

        products.append({
            "ma_vach": ma_vach,
            "name": name,
            "danh_muc": danh_muc,
            "gia_ban": gia_ban,
            "link_anh": link_anh,
            "tinh_trang_kho": tinh_trang,
            "chien_luoc": chien_luoc,
            "uu_tien": uu_tien,
        })

    if not products:
        raise ValueError("Không tìm thấy sản phẩm nào trong file. Kiểm tra file từ dòng 3 trở đi.")

    return products, skipped
