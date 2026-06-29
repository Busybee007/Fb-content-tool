import io
import os
from datetime import date, timedelta

# Load .env nếu có
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
from uuid import uuid4

from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required, login_user, logout_user
from werkzeug.utils import secure_filename

import db
from auth import User, admin_required, login_manager, order_required, page_access_required
from services.excel_import import parse_excel
from services.page_allocator import DAYS, DEFAULT_PER_DAY, allocate_page_week

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "facebook-content-tool-dev-2026")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
app.config["UPLOAD_FOLDER"] = os.path.join(app.root_path, "uploads")
app.config["MEDIA_FOLDER"] = os.environ.get(
    "MEDIA_DIR", os.path.join(app.root_path, "static", "media")
)

ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp"}

login_manager.init_app(app)
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["MEDIA_FOLDER"], exist_ok=True)
db.init_db()

APP_NAME = "FBTool"

WEEKDAY_VI = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]


@app.context_processor
def inject_globals():
    return {"app_name": APP_NAME}


@app.errorhandler(403)
def forbidden(_e):
    return render_template("403.html"), 403


@app.errorhandler(413)
def too_large(_e):
    flash("File quá lớn. Dung lượng tối đa là 50MB.", "danger")
    return redirect(url_for("upload"))


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("pages_list"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user_data = db.check_user_password(username, password)
        if user_data:
            login_user(User(user_data))
            return redirect(request.args.get("next") or url_for("pages_list"))
        flash("Tên đăng nhập hoặc mật khẩu không đúng.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ── Upload (admin only) ───────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return redirect(url_for("pages_list"))


@app.route("/upload", methods=["GET", "POST"])
@admin_required
def upload():
    preview = db.get_products(limit=10) if db.count_products() > 0 else []
    total = db.count_products()
    categories = db.get_categories()

    if request.method == "POST":
        file = request.files.get("excel_file")
        if not file or not file.filename:
            flash("Vui lòng chọn file Excel.", "danger")
            return redirect(url_for("upload"))
        if not file.filename.lower().endswith((".xlsx", ".xlsm")):
            flash("Chỉ hỗ trợ file .xlsx hoặc .xlsm.", "danger")
            return redirect(url_for("upload"))

        file_bytes = file.read()
        try:
            products, _ = parse_excel(file_bytes)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("upload"))
        except Exception:
            flash("Không đọc được file Excel. Kiểm tra định dạng file.", "danger")
            return redirect(url_for("upload"))

        filename = secure_filename(file.filename)
        save_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uuid4().hex}_{filename}")
        with open(save_path, "wb") as f:
            f.write(file_bytes)

        inserted, updated, missing = db.upsert_products(products)
        parts = []
        if inserted: parts.append(f"thêm mới {inserted} sản phẩm")
        if updated:  parts.append(f"cập nhật {updated} sản phẩm")
        flash(f"Import thành công: {', '.join(parts) or 'không có thay đổi'}.", "success")
        if missing:
            flash(
                f"Lưu ý: {missing} sản phẩm trong hệ thống không có trong file vừa upload "
                f"(có thể bị sót). Kiểm tra lại nếu cần.",
                "warning",
            )
        return redirect(url_for("upload"))

    return render_template("upload.html", preview=preview, total=total,
                           category_count=len(categories))


# ── Template download ─────────────────────────────────────────────────────────

@app.route("/download-template")
@login_required
def download_template():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    # ── Sheet 1: Products ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Products"

    headers = ["ma_vach", "ten_san_pham", "danh_muc", "gia_ban",
               "link_anh", "tinh_trang_kho", "chien_luoc", "uu_tien"]

    blue_fill = PatternFill(start_color="1877F2", end_color="1877F2", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    note_font = Font(italic=True, color="856404")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = blue_fill
        c.font = white_bold
        c.alignment = Alignment(horizontal="center")

    notes = [
        "Mã vạch duy nhất (từ Nhanh.vn)",
        "Tên sản phẩm đầy đủ",
        "Danh mục (VD: Áo, Quần, Váy...)",
        "Giá bán — chỉ nhập số",
        "URL ảnh sản phẩm",
        "sẵn hoặc order",
        "mass / thanh lý / mở bán",
        "1 (cao) / 2 (tb) / 3 (thấp)",
    ]
    for col, note in enumerate(notes, 1):
        c = ws.cell(row=2, column=col, value=note)
        c.fill = note_fill
        c.font = note_font

    samples = [
        ("SP001", "Áo thun nam basic trắng size L",    "Áo",   150000, "https://example.com/ao1.jpg",   "sẵn",   "mass",     1),
        ("SP002", "Quần jean nữ slim fit xanh size M",  "Quần", 280000, "https://example.com/quan1.jpg", "sẵn",   "mass",     2),
        ("SP003", "Váy hoa mùa hè size M",              "Váy",  320000, "https://example.com/vay1.jpg",  "sẵn",   "mở bán",   1),
        ("SP004", "Áo khoác dù nam màu đen size L",     "Áo",   450000, "https://example.com/ao2.jpg",   "order", "order",    2),
        ("SP005", "Quần short nữ kaki size S",           "Quần", 120000, "https://example.com/quan2.jpg", "sẵn",   "thanh lý", 3),
        ("SP006", "Áo sơ mi nữ trắng công sở size M",   "Áo",   200000, "https://example.com/ao3.jpg",   "sẵn",   "mass",     2),
        ("SP007", "Đầm dự tiệc đỏ size M",              "Váy",  550000, "https://example.com/dam1.jpg",  "order", "mở bán",   1),
        ("SP008", "Quần tây nam đen slim size L",        "Quần", 380000, "https://example.com/quan3.jpg", "sẵn",   "thanh lý", 3),
    ]
    for row_idx, row in enumerate(samples, 3):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)

    for col, width in enumerate([15, 42, 16, 14, 38, 18, 14, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # ── Sheet 2: Hướng dẫn ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Hướng dẫn")

    guide = [
        ("Cột",            "Mô tả",                          "Giá trị hợp lệ",                                                    "Ví dụ"),
        ("ma_vach",        "Mã vạch từ Nhanh.vn",            "Chuỗi ký tự duy nhất, không trùng",                                 "SP001"),
        ("ten_san_pham",   "Tên đầy đủ của sản phẩm",        "Chuỗi ký tự bất kỳ",                                               "Áo thun nam basic"),
        ("danh_muc",       "Danh mục sản phẩm",              "Tự định nghĩa (VD: Áo, Quần, Váy, Phụ kiện...)",                   "Áo"),
        ("gia_ban",        "Giá bán lẻ (VNĐ)",               "Số nguyên dương, không có ký tự đặc biệt",                          "150000"),
        ("link_anh",       "URL ảnh sản phẩm",               "URL hợp lệ bắt đầu bằng https://",                                 "https://cdn.example.com/img.jpg"),
        ("tinh_trang_kho", "Tình trạng tồn kho",             "sẵn — có hàng giao ngay\norder — phải đặt hàng trước",             "sẵn"),
        ("chien_luoc",     "Chiến lược bán hàng",            "mass — bán đại trà\nthanh lý — cần giải phóng hàng\nmở bán — hàng mới ra mắt", "mass"),
        ("uu_tien",        "Mức độ ưu tiên hiển thị",        "1 — ưu tiên cao nhất\n2 — ưu tiên trung bình\n3 — ưu tiên thấp",  "2"),
    ]

    for row_idx, row in enumerate(guide, 1):
        for col, val in enumerate(row, 1):
            c = ws2.cell(row=row_idx, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx == 1:
                c.fill = blue_fill
                c.font = white_bold
                c.alignment = Alignment(horizontal="center", wrap_text=True)

    for col, width in enumerate([20, 32, 52, 35], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = width
    ws2.row_dimensions[1].height = 22
    for r in range(2, len(guide) + 1):
        ws2.row_dimensions[r].height = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="bang_san_pham_mau.xlsx",
    )


# ── Pages list ────────────────────────────────────────────────────────────────

@app.route("/pages")
@login_required
def pages_list():
    if current_user.role == "admin":
        pages = db.get_fanpages()
    else:
        pages = db.get_user_pages(current_user.id)
    return render_template("pages_list.html", pages=pages)


# ── Per-page: setup ───────────────────────────────────────────────────────────

@app.route("/pages/<int:page_id>/setup", methods=["GET", "POST"])
@page_access_required
def page_setup(page_id):
    fanpage = db.get_fanpage(page_id)
    if not fanpage:
        abort(404)

    products = db.get_all_products()
    if not products:
        flash("Chưa có sản phẩm. Admin cần upload Excel trước.", "warning")
        return redirect(url_for("pages_list"))

    categories = db.get_categories()
    ratios = db.get_page_ratios(page_id)
    filter_cfg = db.get_page_setting(page_id, "filter", {"min_price": 0, "max_price": 999_999_999})
    per_day = db.get_page_setting(page_id, "per_day", DEFAULT_PER_DAY)

    # Ensure all categories have a default
    for cat in categories:
        if cat not in ratios["category"]:
            ratios["category"][cat] = round(100 / len(categories), 1) if categories else 0

    if request.method == "POST":
        min_price = float(request.form.get("min_price") or 0)
        max_price = float(request.form.get("max_price") or 999_999_999)
        per_day_val = int(request.form.get("per_day") or DEFAULT_PER_DAY)
        allow_repeat = request.form.get("allow_repeat") == "on"
        week_start_str = request.form.get("week_start") or _current_monday()

        if min_price > max_price:
            flash("Giá tối thiểu không được lớn hơn giá tối đa.", "danger")
            return redirect(url_for("page_setup", page_id=page_id))

        type_keys = ["mass", "thanh ly", "mo ban", "order"]
        new_type = {}
        for k in type_keys:
            try:
                new_type[k] = float(request.form.get(f"type_{k.replace(' ', '_')}", 0))
            except ValueError:
                new_type[k] = 0.0

        if abs(sum(new_type.values()) - 100) > 0.5:
            flash(f"Tổng tỷ lệ loại hàng phải bằng 100% (hiện tại: {sum(new_type.values()):.1f}%).", "danger")
            return redirect(url_for("page_setup", page_id=page_id))

        new_cat = {}
        for cat in categories:
            try:
                new_cat[cat] = float(request.form.get(f"ratio_{cat}", 0))
            except ValueError:
                new_cat[cat] = 0.0

        if abs(sum(new_cat.values()) - 100) > 0.5:
            flash(f"Tổng tỷ lệ danh mục phải bằng 100% (hiện tại: {sum(new_cat.values()):.1f}%).", "danger")
            return redirect(url_for("page_setup", page_id=page_id))

        db.save_page_ratios(page_id, new_type, new_cat)
        db.set_page_setting(page_id, "filter", {"min_price": min_price, "max_price": max_price})
        db.set_page_setting(page_id, "per_day", per_day_val)

        # Bảo vệ content: nếu tuần đã có caption thì yêu cầu xác nhận
        if db.has_week_content(page_id, week_start_str) and not request.form.get("confirm_overwrite"):
            flash("Tuần này đã có content đã viết. Tích vào ô 'Xác nhận tạo lại' bên dưới để tiếp tục (toàn bộ content tuần này sẽ bị xóa).", "warning")
            return redirect(url_for("page_setup", page_id=page_id))

        # Generate 7-day schedule
        start = date.fromisoformat(week_start_str)
        slots = allocate_page_week(products, new_type, new_cat, min_price, max_price,
                                   start, per_day_val, allow_repeat=allow_repeat)
        db.clear_schedule(page_id, week_start_str)
        db.create_schedule(page_id, slots)

        flash(f"Đã tạo lịch {DAYS} ngày — {len(slots)} sản phẩm từ {start.strftime('%d/%m/%Y')}.", "success")
        return redirect(url_for("page_schedule", page_id=page_id, week_start=week_start_str))

    return render_template(
        "page_setup.html",
        fanpage=fanpage,
        categories=categories,
        ratios=ratios,
        filter_cfg=filter_cfg,
        per_day=per_day,
        current_monday=_current_monday(),
        product_count=len(products),
    )


# ── Per-page: schedule view ───────────────────────────────────────────────────

@app.route("/pages/<int:page_id>/schedule")
@page_access_required
def page_schedule(page_id):
    fanpage = db.get_fanpage(page_id)
    if not fanpage:
        abort(404)

    week_start = request.args.get("week_start") or _current_monday()
    schedule = db.get_schedule(page_id, week_start)
    weeks = db.get_schedule_weeks(page_id)

    start = date.fromisoformat(week_start)
    date_labels = [
        {"date": (start + timedelta(days=i)).isoformat(),
         "label": WEEKDAY_VI[i],
         "display": (start + timedelta(days=i)).strftime("%d/%m")}
        for i in range(7)
    ]
    total_slots = sum(len(v) for v in schedule.values())

    all_products = db.get_all_products()
    return render_template(
        "page_schedule.html",
        fanpage=fanpage,
        schedule=schedule,
        date_labels=date_labels,
        week_start=week_start,
        weeks=weeks,
        total_slots=total_slots,
        all_products=all_products,
    )


@app.route("/pages/<int:page_id>/schedule/swap", methods=["POST"])
@page_access_required
def schedule_swap(page_id):
    if request.is_json:
        data = request.get_json(silent=True) or {}
        slot_a = data.get("slot_a")
        slot_b = data.get("slot_b")
        week_start = data.get("week_start") or _current_monday()
        if not slot_a or not slot_b or slot_a == slot_b:
            return jsonify({"ok": False, "error": "Chọn đủ 2 sản phẩm khác nhau"}), 400
        if not db.swap_schedule_slots(slot_a, slot_b):
            return jsonify({"ok": False, "error": "Hoán đổi thất bại"}), 500
        return jsonify({"ok": True})

    slot_a = request.form.get("slot_a", type=int)
    slot_b = request.form.get("slot_b", type=int)
    week_start = request.form.get("week_start") or _current_monday()
    if not slot_a or not slot_b or slot_a == slot_b:
        flash("Chọn đủ 2 sản phẩm khác nhau để hoán đổi.", "danger")
    elif not db.swap_schedule_slots(slot_a, slot_b):
        flash("Hoán đổi thất bại.", "danger")
    else:
        flash("Đã hoán đổi vị trí 2 sản phẩm.", "success")
    return redirect(url_for("page_schedule", page_id=page_id, week_start=week_start))


# ── Per-page: content (captions) ─────────────────────────────────────────────

@app.route("/pages/<int:page_id>/schedule/week-status")
@page_access_required
def schedule_week_status(page_id):
    week_start = request.args.get("week_start") or _current_monday()
    return jsonify({"has_content": db.has_week_content(page_id, week_start)})


@app.route("/pages/<int:page_id>/schedule/replace", methods=["POST"])
@page_access_required
def schedule_replace(page_id):
    slot_id = request.json.get("slot_id") if request.is_json else request.form.get("slot_id", type=int)
    product_id = request.json.get("product_id") if request.is_json else request.form.get("product_id", type=int)
    if not slot_id or not product_id:
        return jsonify({"ok": False, "error": "Thiếu tham số"}), 400
    db.replace_slot_product(slot_id, product_id)
    product = db.get_product(product_id)
    return jsonify({"ok": True, "product": product})


@app.route("/pages/<int:page_id>/content", methods=["GET", "POST"])
@page_access_required
def page_content(page_id):
    fanpage = db.get_fanpage(page_id)
    if not fanpage:
        abort(404)

    week_start = request.args.get("week_start") or _current_monday()
    schedule = db.get_schedule(page_id, week_start)
    all_slots = [s for slots in schedule.values() for s in slots]

    if not all_slots:
        flash("Chưa có lịch cho tuần này. Hãy thiết lập và tạo lịch trước.", "warning")
        return redirect(url_for("page_setup", page_id=page_id))

    if request.method == "POST":
        slot_id = request.form.get("slot_id", type=int)
        caption = request.form.get("caption", "").strip()
        if slot_id:
            db.update_slot_caption(slot_id, caption)
            flash("Đã lưu caption.", "success")
        return redirect(url_for("page_content", page_id=page_id, week_start=week_start,
                                slot_id=slot_id))

    active_slot_id = request.args.get("slot_id", type=int)
    if not active_slot_id:
        active_slot_id = all_slots[0]["id"]
    active_slot = next((s for s in all_slots if s["id"] == active_slot_id), all_slots[0])

    return render_template(
        "page_content.html",
        fanpage=fanpage,
        all_slots=all_slots,
        active_slot=active_slot,
        week_start=week_start,
    )


# ── Per-page: media upload ────────────────────────────────────────────────────

@app.route("/pages/<int:page_id>/content/media", methods=["POST"])
@page_access_required
def slot_media(page_id):
    slot_id = request.form.get("slot_id", type=int)
    action = request.form.get("action")
    week_start = request.form.get("week_start") or _current_monday()
    redirect_url = url_for("page_content", page_id=page_id,
                           week_start=week_start, slot_id=slot_id)

    if not slot_id:
        abort(400)

    if action == "delete":
        # Delete old file if image
        slot = db.get_slot(slot_id)
        if slot and slot.get("media_type") == "image" and slot.get("media_value"):
            old_path = os.path.join(app.root_path, "static", slot["media_value"])
            if os.path.exists(old_path):
                os.remove(old_path)
        db.clear_slot_media(slot_id)
        flash("Đã xoá media.", "success")
        return redirect(redirect_url)

    if action == "upload_image":
        file = request.files.get("image_file")
        if not file or not file.filename:
            flash("Chưa chọn file ảnh.", "danger")
            return redirect(redirect_url)
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ALLOWED_IMAGE_EXT:
            flash("Chỉ hỗ trợ JPG, PNG, GIF, WEBP.", "danger")
            return redirect(redirect_url)
        # Delete old image if exists
        slot = db.get_slot(slot_id)
        if slot and slot.get("media_type") == "image" and slot.get("media_value"):
            old_path = os.path.join(app.root_path, "static", slot["media_value"])
            if os.path.exists(old_path):
                os.remove(old_path)
        filename = f"slot_{slot_id}_{uuid4().hex}{ext}"
        file.save(os.path.join(app.config["MEDIA_FOLDER"], filename))
        db.update_slot_media(slot_id, "image", f"media/{filename}")
        flash("Đã tải ảnh lên.", "success")
        return redirect(redirect_url)

    if action == "set_video":
        video_url = request.form.get("video_url", "").strip()
        if not video_url:
            flash("Chưa nhập link video.", "danger")
            return redirect(redirect_url)
        db.update_slot_media(slot_id, "video", video_url)
        flash("Đã lưu link video.", "success")
        return redirect(redirect_url)

    abort(400)


# ── Admin: users ──────────────────────────────────────────────────────────────

@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def admin_users():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "create":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()
            role = request.form.get("role", "employee")
            if not username or not password:
                flash("Cần nhập đầy đủ tên đăng nhập và mật khẩu.", "danger")
            elif db.get_user_by_username(username):
                flash(f"Tên đăng nhập '{username}' đã tồn tại.", "danger")
            else:
                db.create_user(username, password, role)
                flash(f"Đã tạo tài khoản '{username}'.", "success")

        elif action == "reset_password":
            user_id = request.form.get("user_id", type=int)
            new_pw = request.form.get("new_password", "").strip()
            if user_id and new_pw:
                db.update_user_password(user_id, new_pw)
                flash("Đã đổi mật khẩu.", "success")

        elif action == "toggle_order":
            user_id = request.form.get("user_id", type=int)
            if user_id:
                target = db.get_user_by_id(user_id)
                if target:
                    new_val = not bool(target.get("can_order", 0))
                    db.set_user_can_order(user_id, new_val)
                    state = "bật" if new_val else "tắt"
                    flash(f"Đã {state} quyền Order Tool cho '{target['username']}'.", "success")

        elif action == "delete":
            user_id = request.form.get("user_id", type=int)
            if user_id == current_user.id:
                flash("Không thể xoá tài khoản đang đăng nhập.", "danger")
            elif user_id:
                db.delete_user(user_id)
                flash("Đã xoá tài khoản.", "success")

        return redirect(url_for("admin_users"))

    users = db.get_all_users()
    return render_template("admin_users.html", users=users)


# ── Admin: pages (fanpages) ───────────────────────────────────────────────────

@app.route("/admin/pages", methods=["GET", "POST"])
@admin_required
def admin_pages():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "create":
            name = request.form.get("name", "").strip()
            fb_url = request.form.get("fb_url", "").strip()
            description = request.form.get("description", "").strip()
            if not name:
                flash("Tên page không được để trống.", "danger")
            else:
                db.create_fanpage(name, fb_url, description)
                flash(f"Đã tạo page '{name}'.", "success")

        elif action == "update":
            page_id = request.form.get("page_id", type=int)
            name = request.form.get("name", "").strip()
            fb_url = request.form.get("fb_url", "").strip()
            description = request.form.get("description", "").strip()
            if page_id and name:
                db.update_fanpage(page_id, name, fb_url, description)
                flash("Đã cập nhật thông tin page.", "success")

        elif action == "delete":
            page_id = request.form.get("page_id", type=int)
            if page_id:
                db.delete_fanpage(page_id)
                flash("Đã xoá page.", "success")

        elif action == "assign":
            page_id = request.form.get("page_id", type=int)
            user_id = request.form.get("user_id", type=int)
            if page_id and user_id:
                db.assign_user_to_page(user_id, page_id)
                flash("Đã phân quyền.", "success")

        elif action == "unassign":
            page_id = request.form.get("page_id", type=int)
            user_id = request.form.get("user_id", type=int)
            if page_id and user_id:
                db.remove_user_from_page(user_id, page_id)
                flash("Đã thu hồi quyền.", "success")

        return redirect(url_for("admin_pages"))

    pages = db.get_fanpages()
    all_users = db.get_all_users()
    page_users = {p["id"]: db.get_page_users(p["id"]) for p in pages}
    # Pre-compute employees not yet assigned per page (for the assign dropdown)
    employee_users = [u for u in all_users if u["role"] != "admin"]
    page_unassigned = {
        p["id"]: [u for u in employee_users
                  if u["id"] not in {pu["id"] for pu in page_users[p["id"]]}]
        for p in pages
    }
    return render_template("admin_pages.html", pages=pages, all_users=all_users,
                           page_users=page_users, page_unassigned=page_unassigned)


# ── Admin: product management ────────────────────────────────────────────────

@app.route("/products")
@admin_required
def products_list():
    q       = request.args.get("q", "").strip()
    cat_f   = request.args.get("cat", "").strip()
    type_f  = request.args.get("type", "").strip()
    products = db.get_all_products()
    if q:
        ql = q.lower()
        products = [p for p in products
                    if ql in (p.get("name") or "").lower()
                    or ql in (p.get("ma_vach") or "").lower()]
    if cat_f:
        products = [p for p in products if p.get("danh_muc") == cat_f]
    if type_f:
        products = [p for p in products if p.get("chien_luoc") == type_f]
    categories = db.get_categories()
    return render_template("products.html", products=products,
                           categories=categories, q=q, cat_f=cat_f, type_f=type_f)


@app.route("/products/add", methods=["POST"])
@admin_required
def product_add():
    ma_vach       = request.form.get("ma_vach", "").strip()
    name          = request.form.get("name", "").strip()
    link_anh      = request.form.get("link_anh", "").strip()
    gia_ban       = request.form.get("gia_ban", "0")
    danh_muc      = request.form.get("danh_muc", "").strip()
    tinh_trang_kho= request.form.get("tinh_trang_kho", "san")
    chien_luoc    = request.form.get("chien_luoc", "mass")
    uu_tien       = request.form.get("uu_tien", "2")
    if not ma_vach or not name:
        flash("Mã vạch và tên sản phẩm không được để trống.", "danger")
    else:
        _, err = db.add_product(ma_vach, name, link_anh, gia_ban,
                                danh_muc, tinh_trang_kho, chien_luoc, uu_tien)
        if err:
            flash(err, "danger")
        else:
            flash(f"Đã thêm sản phẩm '{name}' ({ma_vach}).", "success")
    return redirect(url_for("products_list"))


@app.route("/products/<int:product_id>/edit", methods=["POST"])
@admin_required
def product_edit(product_id):
    p = db.get_product(product_id)
    if not p:
        flash("Không tìm thấy sản phẩm.", "danger")
        return redirect(url_for("products_list"))
    # Only update fields present in form; fall back to existing values
    chien_luoc     = request.form.get("chien_luoc", p["chien_luoc"])
    tinh_trang_kho = request.form.get("tinh_trang_kho", p["tinh_trang_kho"])
    uu_tien        = request.form.get("uu_tien", p["uu_tien"])
    name           = request.form.get("name", p["name"]) or p["name"]
    link_anh       = request.form.get("link_anh", p["link_anh"] or "")
    gia_ban        = request.form.get("gia_ban", p["gia_ban"])
    danh_muc       = request.form.get("danh_muc", p["danh_muc"] or "")
    db.update_product(product_id, name, link_anh, gia_ban,
                      danh_muc, tinh_trang_kho, chien_luoc, uu_tien)
    flash("Đã cập nhật sản phẩm.", "success")
    return redirect(url_for("products_list"))


@app.route("/products/<int:product_id>/delete", methods=["POST"])
@admin_required
def product_delete(product_id):
    p = db.get_product(product_id)
    if not p:
        flash("Không tìm thấy sản phẩm.", "danger")
        return redirect(url_for("products_list"))
    slot_count = db.count_product_slots(product_id)
    confirmed = request.form.get("confirm_delete") == "1"
    if slot_count > 0 and not confirmed:
        flash(
            f"Sản phẩm '{p['name']}' đang có trong {slot_count} slot lịch. "
            f"Xác nhận xóa sẽ xóa luôn các slot đó.",
            "warning",
        )
        return redirect(url_for("products_list"))
    db.delete_product(product_id)
    flash(f"Đã xóa sản phẩm '{p['name']}'.", "success")
    return redirect(url_for("products_list"))


# ── Content library ───────────────────────────────────────────────────────────

@app.route("/help")
@login_required
def help_page():
    return render_template("help.html")


@app.route("/content-library")
@login_required
def content_library():
    products = db.get_all_library_products()
    return render_template("content_library.html", products=products)


@app.route("/api/library/<path:ma_vach>")
@login_required
def api_library_list(ma_vach):
    return jsonify(db.get_library_versions(ma_vach))


@app.route("/api/library", methods=["POST"])
@login_required
def api_library_create():
    ma_vach = request.form.get("ma_vach", "").strip()
    if not ma_vach:
        return jsonify({"ok": False, "error": "Thiếu mã vạch"}), 400
    title      = request.form.get("title", "").strip()
    caption    = request.form.get("caption", "").strip()
    post_time  = request.form.get("post_time", "").strip()
    media_type = request.form.get("media_type", "").strip()
    media_value= request.form.get("media_value", "").strip()
    file = request.files.get("image_file")
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ALLOWED_IMAGE_EXT:
            return jsonify({"ok": False, "error": "Chỉ hỗ trợ JPG, PNG, GIF, WEBP"}), 400
        filename = f"lib_{ma_vach.replace('/', '_')}_{uuid4().hex}{ext}"
        file.save(os.path.join(app.config["MEDIA_FOLDER"], filename))
        media_type  = "image"
        media_value = f"media/{filename}"
    drive_img = request.form.get("drive_image_url", "").strip()
    vid = request.form.get("video_url", "").strip()
    if drive_img and not file:
        media_type  = "image"
        media_value = drive_img
    elif vid and not file:
        media_type  = "video"
        media_value = vid
    version_id = db.create_library_version(
        ma_vach, title, caption, media_type, media_value, post_time, current_user.id
    )
    return jsonify({"ok": True, "id": version_id,
                    "media_type": media_type, "media_value": media_value})


@app.route("/api/library/version/<int:version_id>")
@login_required
def api_library_get(version_id):
    ver = db.get_library_version(version_id)
    if not ver:
        return jsonify({"ok": False}), 404
    return jsonify(ver)


@app.route("/api/library/<int:version_id>", methods=["POST"])
@login_required
def api_library_update(version_id):
    ver = db.get_library_version(version_id)
    if not ver:
        return jsonify({"ok": False, "error": "Không tìm thấy"}), 404
    title      = request.form.get("title", "").strip()
    caption    = request.form.get("caption", "").strip()
    post_time  = request.form.get("post_time", "").strip()
    media_type = ver["media_type"]
    media_value= ver["media_value"]
    file = request.files.get("image_file")
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ALLOWED_IMAGE_EXT:
            return jsonify({"ok": False, "error": "Chỉ hỗ trợ JPG, PNG, GIF, WEBP"}), 400
        filename = f"lib_{ver['ma_vach'].replace('/', '_')}_{uuid4().hex}{ext}"
        file.save(os.path.join(app.config["MEDIA_FOLDER"], filename))
        media_type  = "image"
        media_value = f"media/{filename}"
    elif request.form.get("remove_media") == "1":
        media_type  = ""
        media_value = ""
    elif request.form.get("drive_image_url", "").strip():
        media_type  = "image"
        media_value = request.form.get("drive_image_url", "").strip()
    elif request.form.get("video_url", "").strip():
        media_type  = "video"
        media_value = request.form.get("video_url", "").strip()
    db.update_library_version(version_id, title, caption, media_type, media_value, post_time)
    return jsonify({"ok": True, "media_type": media_type, "media_value": media_value})


@app.route("/api/library/<int:version_id>/delete", methods=["POST"])
@login_required
def api_library_delete(version_id):
    old = db.delete_library_version(version_id)
    if old and old.get("media_type") == "image" and old.get("media_value"):
        path = os.path.join(app.root_path, "static", old["media_value"])
        if os.path.exists(path):
            os.remove(path)
    return jsonify({"ok": True})


@app.route("/api/library/<int:version_id>/apply", methods=["POST"])
@login_required
def api_library_apply(version_id):
    """Apply a library version as the active content for a product barcode."""
    ver = db.get_library_version(version_id)
    if not ver:
        return jsonify({"ok": False, "error": "Không tìm thấy"}), 404
    db.save_product_content(
        ver["ma_vach"], ver["caption"], ver["media_type"], ver["media_value"], ver["post_time"]
    )
    return jsonify({"ok": True, "version": ver})


# ── Drive photos API (tất cả nhân viên) ──────────────────────────────────────

@app.route("/api/drive/status")
@login_required
def api_drive_status():
    from services import gdrive
    return jsonify({"configured": gdrive.is_configured()})


@app.route("/api/drive/photos/<path:ma_vach>")
@login_required
def api_drive_list(ma_vach):
    from services import gdrive
    if not gdrive.is_configured():
        return jsonify({"photos": [], "folder": None, "configured": False})
    try:
        photos, folder = gdrive.list_photos_by_barcode(ma_vach)
        return jsonify({"photos": photos, "folder": folder, "configured": True})
    except Exception as e:
        return jsonify({"photos": [], "folder": None, "configured": True, "error": str(e)})


# ── Product photos API ───────────────────────────────────────────────────────

@app.route("/api/product-photos/<path:ma_vach>")
@login_required
def api_get_product_photos(ma_vach):
    photos = db.get_product_photos(ma_vach)
    for ph in photos:
        ph["url"] = "/" + ph["filepath"].replace("\\", "/")
    return jsonify(photos)


@app.route("/api/product-photos", methods=["POST"])
@login_required
def api_upload_product_photo():
    ma_vach = request.form.get("ma_vach", "").strip()
    if not ma_vach:
        return jsonify({"ok": False, "error": "Thiếu mã vạch"}), 400
    file = request.files.get("photo")
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "Thiếu file"}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_IMAGE_EXT:
        return jsonify({"ok": False, "error": "Chỉ hỗ trợ JPG, PNG, GIF, WEBP"}), 400

    safe_bc = ma_vach.replace("/", "_").replace("\\", "_")
    photo_dir = os.path.join(app.config["MEDIA_FOLDER"], "photos", safe_bc)
    os.makedirs(photo_dir, exist_ok=True)

    filename  = f"{uuid4().hex}{ext}"
    full_path = os.path.join(photo_dir, filename)
    file.save(full_path)

    # filepath relative to static/  so we can build URL
    rel = os.path.relpath(full_path, os.path.join(app.root_path, "static")).replace("\\", "/")
    photo_id = db.add_product_photo(ma_vach, secure_filename(file.filename), f"static/{rel}", current_user.id)
    return jsonify({"ok": True, "id": photo_id, "url": f"/static/{rel}", "filename": secure_filename(file.filename)})


@app.route("/api/product-photos/<int:photo_id>/delete", methods=["POST"])
@login_required
def api_delete_product_photo(photo_id):
    ph = db.delete_product_photo(photo_id)
    if ph and ph.get("filepath"):
        full = os.path.join(app.root_path, ph["filepath"])
        if os.path.exists(full):
            os.remove(full)
    return jsonify({"ok": True})


# ── Product content API ───────────────────────────────────────────────────────

@app.route("/api/product-content/<path:ma_vach>")
@login_required
def api_get_product_content(ma_vach):
    content = db.get_product_content(ma_vach)
    return jsonify(content or {})


@app.route("/api/product-content", methods=["POST"])
@login_required
def api_save_product_content():
    ma_vach = request.form.get("ma_vach", "").strip()
    if not ma_vach:
        return jsonify({"ok": False, "error": "Thiếu mã vạch"}), 400

    caption = request.form.get("caption", "").strip()
    post_time = request.form.get("post_time", "").strip()
    media_type = request.form.get("media_type", "").strip()
    media_value = request.form.get("media_value", "").strip()

    file = request.files.get("image_file")
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ALLOWED_IMAGE_EXT:
            return jsonify({"ok": False, "error": "Chỉ hỗ trợ JPG, PNG, GIF, WEBP"}), 400
        filename = f"pc_{ma_vach.replace('/', '_')}_{uuid4().hex}{ext}"
        file.save(os.path.join(app.config["MEDIA_FOLDER"], filename))
        media_type = "image"
        media_value = f"media/{filename}"

    db.save_product_content(ma_vach, caption, media_type, media_value, post_time)
    return jsonify({"ok": True, "media_type": media_type, "media_value": media_value})


# ── Helpers ───────────────────────────────────────────────────────────────────

def _current_monday():
    today = date.today()
    return (today - timedelta(days=today.weekday())).isoformat()


# ── Order Tool ────────────────────────────────────────────────────────────────

@app.route("/order-tool", methods=["GET", "POST"])
@order_required
def order_tool():
    supplier_count = db.count_supplier_map()

    if request.method == "POST":
        action = request.form.get("action")

        # ── Cập nhật bảng NCC từ file Products ───────────────────────────────
        if action in ("process", "update_supplier"):
            supplier_file = request.files.get("supplier_file")
            if supplier_file and supplier_file.filename:
                rows = _parse_excel_cols(
                    supplier_file.read(),
                    col_barcode="Mã vạch",
                    col_name="Tên sản phẩm",
                    col_extra="Nhà cung cấp",
                )
                supplier_rows = [
                    (r["barcode"], r["name"] or "", r["extra"] or "")
                    for r in rows if r["barcode"]
                ]
                saved = db.upsert_supplier_map(supplier_rows)
                supplier_count = db.count_supplier_map()
                if action == "update_supplier":
                    flash(f"Đã cập nhật {saved} sản phẩm vào bảng NCC ({supplier_count} tổng).", "success")
                    return redirect(url_for("order_tool"))

        if action != "process":
            return redirect(url_for("order_tool"))

        # ── Xử lý đơn hàng ───────────────────────────────────────────────────
        order_file = request.files.get("order_file")
        inventory_file = request.files.get("inventory_file")

        if not order_file or not order_file.filename:
            flash("Vui lòng chọn file đơn hàng.", "danger")
            return redirect(url_for("order_tool"))
        if not inventory_file or not inventory_file.filename:
            flash("Vui lòng chọn file tồn kho.", "danger")
            return redirect(url_for("order_tool"))

        # Parse đơn hàng → gộp SL theo barcode
        order_rows = _parse_excel_cols(
            order_file.read(),
            col_barcode="Mã vạch",
            col_name="Sản phẩm",
            col_extra="Số lượng",
        )
        orders = {}
        for r in order_rows:
            bc = str(r["barcode"]).strip() if r["barcode"] else ""
            if not bc or bc == "None":
                continue
            qty = int(r["extra"] or 0)
            if bc not in orders:
                orders[bc] = {"name": r["name"] or "", "qty": 0}
            orders[bc]["qty"] += qty

        # Parse tồn kho → {barcode: available}
        inv_rows = _parse_excel_cols(
            inventory_file.read(),
            col_barcode="Barcode",
            col_name=None,
            col_extra="K.Dụng",
        )
        inventory = {}
        for r in inv_rows:
            bc = str(r["barcode"]).strip() if r["barcode"] else ""
            if bc:
                inventory[bc] = int(r["extra"] or 0)

        # Lấy bảng NCC
        supplier_map = db.get_supplier_map()

        # Tính cần nhập
        results = []
        for bc, info in orders.items():
            needed = info["qty"]
            available = inventory.get(bc, 0)
            to_order = max(0, needed - available)
            if to_order == 0:
                continue
            sup_info = supplier_map.get(bc, {})
            results.append({
                "barcode": bc,
                "name": info["name"],
                "total_ordered": needed,
                "available": available,
                "to_order": to_order,
                "supplier": sup_info.get("supplier") or "Chưa xác định",
            })

        if not results:
            flash("Không có sản phẩm nào cần đặt thêm — tồn kho đủ hết!", "success")
            return redirect(url_for("order_tool"))

        results.sort(key=lambda x: (x["supplier"], x["barcode"]))

        # Xuất Excel
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Đề xuất đặt hàng"

        headers = ["STT", "Barcode", "Tên sản phẩm", "SL đơn hàng", "Tồn KD", "Cần nhập", "Nhà cung cấp"]
        widths  = [6, 18, 48, 14, 10, 12, 28]

        header_fill = PatternFill(start_color="1877F2", end_color="1877F2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        supplier_colors = {}
        palette = ["DBEAFE", "D1FAE5", "FEF9C3", "FCE7F3", "EDE9FE", "FFEDD5", "ECFDF5", "FEF2F2"]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[c.column_letter].width = w

        color_idx = 0
        for stt, r in enumerate(results, 1):
            sup = r["supplier"]
            if sup not in supplier_colors:
                supplier_colors[sup] = palette[color_idx % len(palette)]
                color_idx += 1
            fill = PatternFill(start_color=supplier_colors[sup],
                               end_color=supplier_colors[sup], fill_type="solid")
            row_data = [stt, r["barcode"], r["name"],
                        r["total_ordered"], r["available"], r["to_order"], r["supplier"]]
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=stt + 1, column=col, value=val)
                c.fill = fill
                if col in (4, 5, 6):
                    c.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from datetime import datetime
        fname = f"de_xuat_dat_hang_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template("order_tool.html", supplier_count=supplier_count)


def _parse_excel_cols(file_bytes, col_barcode, col_name, col_extra):
    """Parse Excel, trả về list of {barcode, name, extra} dựa vào tên cột."""
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx_bc   = headers.index(col_barcode) if col_barcode in headers else None
    idx_name = headers.index(col_name)    if col_name and col_name in headers else None
    idx_ext  = headers.index(col_extra)   if col_extra and col_extra in headers else None

    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        result.append({
            "barcode": row[idx_bc]   if idx_bc   is not None else None,
            "name":    row[idx_name] if idx_name is not None else None,
            "extra":   row[idx_ext]  if idx_ext  is not None else None,
        })
    return result


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
